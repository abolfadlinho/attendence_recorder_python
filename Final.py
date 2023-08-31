from openpyxl import load_workbook                                      #Imports the load workbook library from openpyxl
import tkinter as tk
from tkinter import filedialog
import tabula

root = tk.Tk()
root.title('Attendance Recorder')
root.geometry("500x730")
root.configure(background="#006400")
root.iconbitmap('icon.ico')

frame = tk.Frame(root, bg="#006400")
frame.pack(fill=tk.X, side=tk.BOTTOM)

msg = ("Easily upload Zoom's meeting report and check who attended your meeting")

canvas = tk.Canvas(root, bg='#006400', width='490', height='100')                                           #Created Window,Canvas,Title and Frame
canvas.pack()
canvas.create_text(250, 20, fill='#02075d', font='Calibri 25 bold', text="Attendance Recorder")
canvas.create_text(250, 70, fill='white', font='Times 11', justify='center', text=msg)

files = []
pagenumber = []


def submit():
    pagenumber.append(str(E1.get()))
    return pagenumber


L1 = tk.Label(root, text="Page Number:")
L1.pack()
E1 = tk.Entry(root, bd=5)
E1.pack()


MyButton1 = tk.Button(root, text="Submit", width=10, command=submit)
MyButton1.pack()


def addFile():
    filename = filedialog.askopenfilename(initialdir='/', title='Select PDF File',
                                         filetypes=(('executables', '*.pdf'), ('all files', '*.*')))
    filenamewithxlsxextension = (filename[:-3]) + "xlsx"

    df = tabula.read_pdf(filename, pages=pagenumber[-1], multiple_tables=True, lattice=True)[1]
    df.to_excel(filenamewithxlsxextension)

    files.append(filenamewithxlsxextension)
    label = tk.Label(root, text=files[0], bg='gray')
    label.pack(side=tk.BOTTOM)


AddFile = tk.Button(frame, command=addFile, text='Add File', font=70)


def AttendanceRecorder():
    book = load_workbook(str(files[0]))

    sheet = book.active
    #This should be an input list, since it was for my own use then I just entered the list
    MES = ['Abdelrahman Ahmed Elshebiny', 'Ahmed Ashraf', 'Farah Sayed Youssef', 'Faress Farouk', 'Hamza Elewa',
           'Hana Ismail', 'Hana Ramy', 'Hashem Abulhassan', 'jana amr', 'Jana emad', 'Jana emad(Jana)', 'Jana Fouda (Fares FOUDA)', 'Jana Fouda',
           'Layla Swelam', 'Layla Swelam MES', 'Mahmoud Ashraf MES', 'maryam elsheikh', 'Mayar Sherif', 'menna samy',
           'mohamed sultan MES', 'nada elattar', 'Omar Ayman', 'Rana Seif', 'sara tarek el bayoumy MES', 'Salma ahmed','jana emad','Jana emad (Jana)','salma walid (salma walid MES)',
           'Youssef Othman', 'Youssef yasser MES', 'Youssuf Amgad']
    all_data = sheet.iter_rows(min_row=1, max_row=100, min_col=1, max_col=5)

    names = []
    for a, b, c, d, e in all_data:
        if b.value in MES and int(d.value) > 30:
            names.append(b.value)
    number_of_attendees = len(names)
    sentence = tk.Label(root, text=(
                str(number_of_attendees) + " Students on page " + pagenumber[-1]),
                        fg='#000000',
                        bg='#FFFFFF',
                        font='Times 20 italic bold')
    sentence.pack(side=tk.TOP)
    for i in range(len(names)):
        string_of_names = (str(i+1)+'. ' + names[i])
        results = tk.Label(root, text=string_of_names, bg='#02075d', fg='#FFFFFF')
        results.pack(fill=tk.X, side=tk.TOP)
    return names


Run = tk.Button(frame, text="Run", font=30, fg='#000000', command=AttendanceRecorder)
Close = tk.Button(frame, text='Close', command=root.destroy, font=30, bg='#FF0000', fg='#FFFFFF')

frame.columnconfigure(0, weight=1)
frame.columnconfigure(1, weight=1)

AddFile.grid(row=0, column=0, sticky=tk.W+tk.E)
Run.grid(row=0, column=1, sticky=tk.W+tk.E)
Close.grid(row=0, column=2, sticky=tk.W+tk.E)

lbl = tk.Label(root, text="Attach your PDF files to record attendance")
lbl.pack(side=tk.BOTTOM)

root.mainloop()
