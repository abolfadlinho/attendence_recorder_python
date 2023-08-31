from openpyxl import load_workbook  #imports the load workbook library from openpyxl

book = load_workbook(input("Input Excel file name:")+".xlsx")  #creates file called book and loads excel file

sheet = book.active
MES = ['Abdelrahman Ahmed Elshebiny','Ahmed Ashraf','Farah Sayed Youssef','Faress Farouk','Hamza Elewa','Hana Ismail', 'Hana Ramy', 'Hashem Abulhassan','jana amr','Jana Emad Gaber','Jana Fouda (Fares FOUDA)','Layla Swelam','Layla Swelam MES', 'Mahmoud Ashraf MES','maryam elsheikh','Mayar Sherif','mohamed sultan MES','nada elattar','Omar Ayman','Rana Seif','sara tarek el bayoumy MES','Salma ahmed','Youssef Othman','Youssef yasser MES','Youssuf Amgad']
#columns = sheet.columns     #gets columns
rows = sheet.rows       #gets rows

#headers = [ cell.value for cell in next(rows)]

all_data = sheet.iter_rows(min_row=3 , max_row=100 , min_col=1 , max_col=5)

names = []
duration = []
print("Names:", "          Durations:")
for a,b,c,d,e in all_data:
    if a.value in MES and c.value >30:
        names.append(a.value)
        duration.append(c.value)
        print(a.value)#,c.value)

    #print("Name:" ,a.value ,"...Duration:" ,c.value)
number_of_attendees = len(names)
print("\n",number_of_attendees,"Students attended the class")#,"\n", names)
#print(duration)